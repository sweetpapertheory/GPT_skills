#!/usr/bin/env python3

from __future__ import annotations

import math
import re
from collections import Counter
from dataclasses import dataclass
from datetime import date, datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


@dataclass
class SourceSpec:
    label: str
    url: str
    source_type: str


def normalize_space(value: Any) -> str:
    return re.sub(r'\s+', ' ', str(value or '')).strip()


def to_int(value: Any, default: int = 0) -> int:
    try:
        if value is None or value == '':
            return default
        return int(float(value))
    except (TypeError, ValueError):
        return default


def to_iso_datetime(epoch_seconds: Any) -> str:
    ts = to_int(epoch_seconds, default=0)
    if ts <= 0:
        return ''
    return datetime.fromtimestamp(ts, tz=timezone.utc).isoformat()


def to_iso_date(epoch_seconds: Any) -> str:
    ts = to_int(epoch_seconds, default=0)
    if ts <= 0:
        return ''
    return datetime.fromtimestamp(ts, tz=timezone.utc).date().isoformat()


def safe_get_url(img_like: Any) -> str:
    if isinstance(img_like, dict):
        for key in ('url_list', 'url', 'uri'):
            val = img_like.get(key)
            if isinstance(val, list) and val:
                return str(val[0])
            if isinstance(val, str) and val:
                return val
    if isinstance(img_like, list) and img_like:
        return safe_get_url(img_like[0])
    return ''


def iter_nodes(obj: Any):
    if isinstance(obj, dict):
        yield obj
        for v in obj.values():
            yield from iter_nodes(v)
    elif isinstance(obj, list):
        for item in obj:
            yield from iter_nodes(item)


def extract_hashtags(aweme: dict[str, Any], source_label: str) -> str:
    tags: list[str] = []
    seen = set()

    def push(tag: Any) -> None:
        t = normalize_space(tag).lstrip('#')
        if not t:
            return
        ht = f'#{t}'
        if ht in seen:
            return
        seen.add(ht)
        tags.append(ht)

    for item in aweme.get('text_extra') or []:
        if not isinstance(item, dict):
            continue
        push(item.get('hashtag_name') or item.get('hashtag'))

    for item in aweme.get('cha_list') or aweme.get('challenge_list') or []:
        if not isinstance(item, dict):
            continue
        push(item.get('cha_name') or item.get('challenge_name'))

    desc = normalize_space(aweme.get('desc'))
    for h in re.findall(r'#[^#\s@]+', desc):
        push(h)

    if source_label.startswith('keyword:'):
        push(source_label.split(':', 1)[1])

    return ' '.join(tags[:40])


def is_verified(author: dict[str, Any]) -> int:
    if not isinstance(author, dict):
        return 0
    if normalize_space(author.get('custom_verify') or author.get('enterprise_verify_reason')):
        return 1
    vtype = author.get('verification_type')
    return 1 if isinstance(vtype, int) and vtype >= 0 else 0


def user_identifier(author: dict[str, Any]) -> str:
    if not isinstance(author, dict):
        return ''
    for key in ('unique_id', 'short_id', 'sec_uid', 'uid', 'nickname'):
        val = normalize_space(author.get(key))
        if val:
            return val
    return ''


def first_effect_ids(aweme: dict[str, Any]) -> str:
    effect_ids: list[str] = []
    seen = set()
    candidates = []
    for key in ('interaction_stickers', 'video_tag', 'anchor_info', 'music'):
        val = aweme.get(key)
        if val is not None:
            candidates.append(val)

    for node in iter_nodes(candidates):
        if not isinstance(node, dict):
            continue
        for key in ('effect_id', 'sticker_id', 'id'):
            if key not in node:
                continue
            raw = normalize_space(node.get(key))
            if raw.isdigit() and raw not in seen:
                seen.add(raw)
                effect_ids.append(raw)
            if len(effect_ids) >= 20:
                return ','.join(effect_ids)
    return ','.join(effect_ids)


def normalize_duration_seconds(aweme: dict[str, Any]) -> int:
    dur = aweme.get('duration')
    if dur is None and isinstance(aweme.get('video'), dict):
        dur = aweme['video'].get('duration')
    val = to_int(dur, default=0)
    if val > 1000:
        return max(int(round(val / 1000.0)), 0)
    return max(val, 0)


def video_quality(aweme: dict[str, Any]) -> str:
    video = aweme.get('video') if isinstance(aweme.get('video'), dict) else {}
    width = to_int(video.get('width'), default=0)
    height = to_int(video.get('height'), default=0)
    if width > 0 and height > 0:
        return f'{width}x{height}'
    return ''


def compute_influence_score(follower_count: int, avg_likes_per_video: float) -> float:
    return round(0.65 * math.log10(follower_count + 1) + 0.35 * math.log10(avg_likes_per_video + 1), 6)


def map_user_record(author: dict[str, Any], profile: dict[str, Any]) -> dict[str, Any]:
    username = user_identifier(author) or user_identifier(profile)
    display_name = normalize_space(author.get('nickname') or profile.get('nickname') or username)
    bio = normalize_space(profile.get('signature') or author.get('signature'))
    avatar = safe_get_url(profile.get('avatar_medium') or profile.get('avatar_thumb') or author.get('avatar_thumb'))

    follower_count = to_int(profile.get('follower_count') or author.get('follower_count'), default=0)
    following_count = to_int(profile.get('following_count') or author.get('following_count'), default=0)
    user_likes_count = to_int(profile.get('total_favorited') or author.get('total_favorited'), default=0)
    user_video_count = to_int(profile.get('aweme_count') or author.get('aweme_count'), default=0)

    ratio = round(follower_count / max(following_count, 1), 6) if follower_count > 0 else 0.0
    avg_likes = round(user_likes_count / max(user_video_count, 1), 6) if user_likes_count > 0 else 0.0
    influence = compute_influence_score(follower_count, avg_likes)

    return {
        'username': username,
        'display_name': display_name,
        'bio_description': bio,
        'avatar_url': avatar,
        'is_verified': is_verified(profile or author),
        'follower_count': follower_count,
        'following_count': following_count,
        'user_likes_count': user_likes_count,
        'user_video_count': user_video_count,
        'follower_following_ratio': ratio,
        'avg_likes_per_video': avg_likes,
        'influence_score': influence,
    }


def map_video_record(
    aweme: dict[str, Any],
    user: dict[str, Any],
    processing_date: str,
    source_label: str,
) -> dict[str, Any]:
    stats = aweme.get('statistics') if isinstance(aweme.get('statistics'), dict) else {}
    author = aweme.get('author') if isinstance(aweme.get('author'), dict) else {}

    video_id = normalize_space(aweme.get('aweme_id'))
    video_url = normalize_space(aweme.get('share_url')) or f'https://www.douyin.com/video/{video_id}'
    desc = normalize_space(aweme.get('desc'))
    voice_text = normalize_space(aweme.get('video_text') or '')

    like_count = to_int(stats.get('digg_count'), default=0)
    comment_count = to_int(stats.get('comment_count'), default=0)
    share_count = to_int(stats.get('share_count'), default=0)
    view_count = to_int(stats.get('play_count'), default=0)
    fav_count = to_int(stats.get('collect_count'), default=0)

    engagement_den = view_count if view_count > 0 else max(user.get('follower_count', 0), 1)
    engagement_rate = round((like_count + comment_count + share_count) / engagement_den, 6)

    return {
        'processing_date': processing_date,
        'video_id': video_id,
        'username': user.get('username') or user_identifier(author),
        'video_url': video_url,
        'create_time': to_iso_datetime(aweme.get('create_time')),
        'create_date': to_iso_date(aweme.get('create_time')),
        'region_code': normalize_space(aweme.get('region')),
        'video_description': desc,
        'voice_to_text': voice_text,
        'like_count': like_count,
        'comment_count': comment_count,
        'share_count': share_count,
        'view_count': view_count,
        'favorites_count': fav_count,
        'video_duration': normalize_duration_seconds(aweme),
        'music_id': normalize_space((aweme.get('music') or {}).get('id')) if isinstance(aweme.get('music'), dict) else '',
        'playlist_id': normalize_space((aweme.get('mix_info') or {}).get('mix_id')) if isinstance(aweme.get('mix_info'), dict) else '',
        'is_stem_verified': is_verified(author),
        'all_hashtags': extract_hashtags(aweme, source_label),
        'effect_ids': first_effect_ids(aweme),
        'engagement_rate': engagement_rate,
        'video_quality': video_quality(aweme),
        'has_voice_to_text': 1 if voice_text else 0,
        'description_length': len(desc),
        'display_name': user.get('display_name') or normalize_space(author.get('nickname')),
        'bio_description': user.get('bio_description') or normalize_space(author.get('signature')),
        'avatar_url': user.get('avatar_url') or safe_get_url(author.get('avatar_thumb')),
        'is_verified': user.get('is_verified', 0),
        'follower_count': user.get('follower_count', 0),
        'following_count': user.get('following_count', 0),
        'user_likes_count': user.get('user_likes_count', 0),
        'user_video_count': user.get('user_video_count', 0),
        'follower_following_ratio': user.get('follower_following_ratio', 0),
        'avg_likes_per_video': user.get('avg_likes_per_video', 0),
        'influence_score': user.get('influence_score', 0),
    }


def write_rows(ws, headers: list[str], rows: list[dict[str, Any]]) -> None:
    if ws.max_row and ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)
    for ridx, row in enumerate(rows, start=2):
        for cidx, h in enumerate(headers, start=1):
            ws.cell(ridx, cidx).value = row.get(h)


def update_daily_distribution(ws, video_rows: list[dict[str, Any]]) -> None:
    if ws.max_row and ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    counts: Counter[str] = Counter()
    for row in video_rows:
        d = normalize_space(row.get('create_date'))
        if d:
            counts[d] += 1
    weekday = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday']
    for ridx, d in enumerate(sorted(counts), start=2):
        y, m, day = (int(x) for x in d.split('-'))
        rec = {'Date': d, 'Video_Count': counts[d], 'Weekday': weekday[date(y, m, day).weekday()]}
        for cidx, h in enumerate(headers, start=1):
            ws.cell(ridx, cidx).value = rec.get(h)


def update_processing_summary(
    ws,
    sources: list[SourceSpec],
    video_rows: list[dict[str, Any]],
    comment_rows: list[dict[str, Any]],
    user_rows: list[dict[str, Any]],
    started_at: datetime,
    ended_at: datetime,
) -> None:
    metrics = [normalize_space(ws.cell(r, 1).value) for r in range(1, ws.max_row + 1)]
    metric_to_row = {m: i for i, m in enumerate(metrics, start=1) if m}

    dates = sorted({normalize_space(v.get('create_date')) for v in video_rows if normalize_space(v.get('create_date'))})
    date_range = f'{dates[0]} - {dates[-1]}' if dates else 'N/A'
    focus = ' + '.join(src.label for src in sources) if sources else 'N/A'
    process_seconds = max((ended_at - started_at).total_seconds(), 0.0)

    values = {
        'Collection Focus': f'Douyin {focus}',
        'Date Range': date_range,
        'Total English Videos': len(video_rows),
        'Total Comments': len(comment_rows),
        'Total Users': len(user_rows),
        'Processing Time': f'{process_seconds:.1f}s',
    }
    for key, val in values.items():
        row = metric_to_row.get(key)
        if row:
            ws.cell(row, 2).value = val


def write_workbook(
    template_path: Path,
    output_path: Path,
    sources: list[SourceSpec],
    video_rows: list[dict[str, Any]],
    comment_rows: list[dict[str, Any]],
    user_rows: list[dict[str, Any]],
    started_at: datetime,
    ended_at: datetime,
) -> None:
    wb = load_workbook(template_path)

    for sheet_name in ['Video_Data', 'Comment_Data', 'User_Data', 'Daily_Distribution', 'Processing_Summary']:
        if sheet_name not in wb.sheetnames:
            raise ValueError(f'Template missing sheet: {sheet_name}')

    ws_v = wb['Video_Data']
    ws_c = wb['Comment_Data']
    ws_u = wb['User_Data']
    ws_d = wb['Daily_Distribution']
    ws_s = wb['Processing_Summary']

    headers_v = [ws_v.cell(1, c).value for c in range(1, ws_v.max_column + 1)]
    headers_c = [ws_c.cell(1, c).value for c in range(1, ws_c.max_column + 1)]
    headers_u = [ws_u.cell(1, c).value for c in range(1, ws_u.max_column + 1)]

    write_rows(ws_v, headers_v, video_rows)
    write_rows(ws_c, headers_c, comment_rows)
    write_rows(ws_u, headers_u, user_rows)
    update_daily_distribution(ws_d, video_rows)
    update_processing_summary(ws_s, sources, video_rows, comment_rows, user_rows, started_at, ended_at)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
