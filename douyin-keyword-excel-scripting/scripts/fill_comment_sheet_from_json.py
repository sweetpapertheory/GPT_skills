#!/usr/bin/env python3

from __future__ import annotations

import argparse
import json
from datetime import date
from pathlib import Path

from openpyxl import load_workbook


def parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser(description='Fill Comment_Data sheet from crawled comments JSON')
    p.add_argument('--workbook-path', required=True)
    p.add_argument('--comments-json', required=True)
    p.add_argument('--output-path', required=True)
    return p.parse_args()


def load_allowed_video_ids(workbook_path: Path) -> set[str]:
    wb = load_workbook(workbook_path, read_only=True, data_only=True)
    if 'Video_Data' not in wb.sheetnames:
        raise ValueError('Workbook missing Video_Data sheet')

    ws = wb['Video_Data']
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    idx = {h: i for i, h in enumerate(headers)}
    if 'video_id' not in idx:
        raise ValueError('Video_Data missing video_id column')

    allowed: set[str] = set()
    i_vid = idx['video_id']
    for row in ws.iter_rows(min_row=2, values_only=True):
        if i_vid >= len(row):
            continue
        video_id = str(row[i_vid] or '').strip()
        if video_id:
            allowed.add(video_id)
    return allowed


def main() -> None:
    args = parse_args()
    wb_path = Path(args.workbook_path)
    cj_path = Path(args.comments_json)
    out = Path(args.output_path)

    if not wb_path.exists():
        raise FileNotFoundError(f'workbook not found: {wb_path}')
    if not cj_path.exists():
        raise FileNotFoundError(f'comments json not found: {cj_path}')

    payload = json.loads(cj_path.read_text(encoding='utf-8'))
    comments = payload.get('comments') or []
    allowed_video_ids = load_allowed_video_ids(wb_path)

    wb = load_workbook(wb_path)
    if 'Comment_Data' not in wb.sheetnames:
        raise ValueError('Workbook missing Comment_Data sheet')

    ws = wb['Comment_Data']
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]

    if ws.max_row and ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)

    seen = set()
    rows = []
    skipped_not_in_video_data = 0
    for c in comments:
        video_id = str(c.get('video_id') or '').strip()
        if allowed_video_ids and video_id not in allowed_video_ids:
            skipped_not_in_video_data += 1
            continue

        cid = str(c.get('comment_id') or '').strip()
        if not cid:
            continue
        dedup_key = f'{video_id}:{cid}'
        if dedup_key in seen:
            continue
        seen.add(dedup_key)

        text = str(c.get('text') or '').strip()
        reply_count = int(c.get('reply_count') or 0)
        parent = str(c.get('parent_comment_id') or '').strip()

        row = {
            'video_id': video_id,
            'comment_id': cid,
            'create_time': str(c.get('create_time') or '').strip(),
            'like_count': int(c.get('like_count') or 0),
            'reply_count': reply_count,
            'parent_comment_id': parent,
            'text': text,
            'processing_date': date.today().isoformat(),
            'comment_length': len(text),
            'has_replies': 1 if reply_count > 0 else 0,
            'is_reply': 1 if parent else 0,
        }
        rows.append(row)

    for ridx, row in enumerate(rows, start=2):
        for cidx, h in enumerate(headers, start=1):
            ws.cell(ridx, cidx).value = row.get(h)

    if 'Processing_Summary' in wb.sheetnames:
        ss = wb['Processing_Summary']
        for r in range(2, ss.max_row + 1):
            metric = str(ss.cell(r, 1).value or '').strip()
            if metric == 'Total Comments':
                ss.cell(r, 2).value = len(rows)
                break

    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    print(f'output={out}')
    print(f'comments={len(rows)}')
    print(f'skipped_comments_not_in_video_data={skipped_not_in_video_data}')


if __name__ == '__main__':
    main()
