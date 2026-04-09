#!/usr/bin/env python3
import atexit
import argparse
import json
import math
import os
import re
import shutil
import subprocess
import sys
import tempfile
from datetime import datetime, timezone
from pathlib import Path

import numpy as np
import openpyxl
import pandas as pd
from PIL import Image, ImageDraw, ImageFont, ImageOps
from asr_policy import resolve_whisper_cli_model
from douyin_raw_publish import publish_douyin_raw_json
try:
    from faster_whisper import WhisperModel
    _FASTER_WHISPER_IMPORT_ERROR = None
except Exception as exc:
    WhisperModel = None
    _FASTER_WHISPER_IMPORT_ERROR = exc
try:
    from opencc import OpenCC
except Exception:
    OpenCC = None

DEFAULT_URL = "https://www.tiktok.com/@complex_math/video/7244938382814563586"
DEFAULT_TEMPLATE_PATH = Path("/Users/XM/Desktop/Climate_SVP_MCP_Template_transcript_v3.xlsx")
DEFAULT_OUTPUT_ROOT = Path("/Users/XM/Desktop/Data/TT data/Data analysis")
_CC_T2S = OpenCC("t2s") if OpenCC is not None else None
_OUTPUT_VIDEO_SUFFIXES = {".mp4", ".mkv", ".webm", ".mov", ".avi", ".m4v"}


def remove_output_video_files(out_dir):
    if not out_dir.exists():
        return
    for p in out_dir.iterdir():
        if p.is_file() and p.suffix.lower() in _OUTPUT_VIDEO_SUFFIXES:
            try:
                p.unlink()
            except Exception:
                pass


def cleanup_tmp_media(video_id):
    for pattern in (f"tiktok_{video_id}.*", f"tiktok_{video_id}_*"):
        for p in Path("/tmp").glob(pattern):
            try:
                p.unlink()
            except Exception:
                pass


def parse_args():
    parser = argparse.ArgumentParser(
        description="Run TikTok MCP pipeline for one URL."
    )
    parser.add_argument("--url", default=DEFAULT_URL, help="TikTok video URL")
    parser.add_argument(
        "--template-path",
        default=str(DEFAULT_TEMPLATE_PATH),
        help="Path to Climate SVP MCP Excel template",
    )
    parser.add_argument(
        "--output-root",
        default=str(DEFAULT_OUTPUT_ROOT),
        help="Root folder where output folders are created (Douyin_<id>_mcp for Douyin URLs, tiktok_<id>_mcp otherwise)",
    )
    parser.add_argument(
        "--keep-jpg-storyboards",
        action="store_true",
        help="Keep legacy JPG storyboards (default deletes JPG and keeps HD PNG storyboards only)",
    )
    parser.add_argument(
        "--asr-backend",
        choices=["auto", "faster-whisper", "whisper-cli"],
        default="auto",
        help="ASR backend. 'auto' tries faster-whisper first, then falls back to whisper CLI.",
    )
    parser.add_argument(
        "--whisper-cli-model",
        default=None,
        help=(
            "Whisper model name used by whisper CLI fallback. "
            "If omitted, the pipeline uses turbo for shorter audio and may auto-select tiny for long audio."
        ),
    )
    parser.add_argument("--internal-asr-json-output", default=None, help=argparse.SUPPRESS)
    parser.add_argument("--internal-audio-path", default=None, help=argparse.SUPPRESS)
    parser.add_argument("--internal-asr-model", default="small", help=argparse.SUPPRESS)
    return parser.parse_args()


def run(cmd, check=True, capture=True):
    print("+", " ".join(str(x) for x in cmd), flush=True)
    proc = subprocess.run(
        cmd,
        text=True,
        stdout=subprocess.PIPE if capture else None,
        stderr=subprocess.PIPE if capture else None,
    )
    if check and proc.returncode != 0:
        stdout_tail = (proc.stdout or "")[-1500:]
        stderr_tail = (proc.stderr or "")[-1500:]
        raise RuntimeError(
            f"Command failed ({proc.returncode}): {' '.join(str(x) for x in cmd)}\n"
            f"STDOUT tail:\n{stdout_tail}\nSTDERR tail:\n{stderr_tail}"
        )
    return proc


def parse_video_id(url):
    m = re.search(r"/video/(\d+)", url)
    if not m:
        raise ValueError(f"Cannot parse video id from URL: {url}")
    return m.group(1)


def iso_utc_from_ts(ts):
    if ts is None:
        return None
    return datetime.fromtimestamp(float(ts), tz=timezone.utc).replace(microsecond=0).isoformat()


def now_iso_utc():
    return datetime.now(timezone.utc).replace(microsecond=0).isoformat()


def parse_hashtags(text):
    tags = re.findall(r"#([^\s#]+)", text or "")
    seen = set()
    out = []
    for t in tags:
        t = t.strip().strip(".,;:!?)(([]{}\"'")
        if not t:
            continue
        key = t.lower()
        if key not in seen:
            out.append(t)
            seen.add(key)
    return out


def to_simplified_text(value):
    if not isinstance(value, str) or not value:
        return value
    if _CC_T2S is None:
        return value
    try:
        return _CC_T2S.convert(value)
    except Exception:
        return value


def to_simplified_obj(value):
    if isinstance(value, str):
        return to_simplified_text(value)
    if isinstance(value, list):
        return [to_simplified_obj(x) for x in value]
    if isinstance(value, dict):
        return {k: to_simplified_obj(v) for k, v in value.items()}
    return value


def normalize_workbook_to_simplified(wb):
    for ws in wb.worksheets:
        # Iterate only populated cells to avoid scanning entire worksheet bounds.
        cell_store = getattr(ws, "_cells", None)
        if isinstance(cell_store, dict):
            cells = cell_store.values()
        else:
            cells = (
                cell
                for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column)
                for cell in row
            )
        for cell in cells:
            if isinstance(cell.value, str) and not cell.value.startswith("="):
                cell.value = to_simplified_text(cell.value)


def is_douyin_video_url(url):
    return "douyin.com/video/" in (url or "")


def output_folder_name(url, video_id):
    prefix = "Douyin" if is_douyin_video_url(url) else "tiktok"
    return f"{prefix}_{video_id}_mcp"


def summarize_asr_segments(segment_records):
    if segment_records:
        asr_df_raw = pd.DataFrame(segment_records)
        for column in ("id", "start", "end", "text", "avg_logprob", "no_speech_prob"):
            if column not in asr_df_raw.columns:
                asr_df_raw[column] = None
        asr_df_raw["text"] = asr_df_raw["text"].fillna("").map(
            lambda x: to_simplified_text(str(x).strip())
        )
        asr_df_raw["start"] = np.round(pd.to_numeric(asr_df_raw["start"], errors="coerce").fillna(0.0), 3)
        asr_df_raw["end"] = np.round(pd.to_numeric(asr_df_raw["end"], errors="coerce").fillna(0.0), 3)
        asr_segments = asr_df_raw[
            ["id", "start", "end", "text", "avg_logprob", "no_speech_prob"]
        ].to_dict(orient="records")

        text_arr = asr_df_raw["text"].to_numpy(dtype=object, copy=False)
        transcript_full = " ".join(text_arr[text_arr != ""]).strip()

        span_arr = (
            asr_df_raw["end"].to_numpy(dtype=float, copy=False)
            - asr_df_raw["start"].to_numpy(dtype=float, copy=False)
        )
        duration_after_vad = round(float(np.clip(span_arr, 0.0, None).sum()), 3)
    else:
        asr_segments = []
        transcript_full = ""
        duration_after_vad = 0.0

    return asr_segments, transcript_full, duration_after_vad


def _run_internal_faster_whisper(audio_path, asr_model, json_output_path):
    if WhisperModel is None:
        raise RuntimeError(f"faster-whisper import failed: {_FASTER_WHISPER_IMPORT_ERROR}")

    model = WhisperModel(asr_model, device="cpu", compute_type="int8")
    segments_iter, info = model.transcribe(str(audio_path), vad_filter=True)
    segments = list(segments_iter)

    if segments:
        n_segments = len(segments)
        asr_df_raw = pd.DataFrame(
            {
                "id": np.arange(1, n_segments + 1, dtype=int),
                "start": np.round(
                    np.fromiter((float(s.start) for s in segments), dtype=float, count=n_segments),
                    3,
                ),
                "end": np.round(
                    np.fromiter((float(s.end) for s in segments), dtype=float, count=n_segments),
                    3,
                ),
                "text": np.fromiter(
                    (to_simplified_text((s.text or "").strip()) for s in segments),
                    dtype=object,
                    count=n_segments,
                ),
                "avg_logprob": np.fromiter(
                    (None if s.avg_logprob is None else float(s.avg_logprob) for s in segments),
                    dtype=object,
                    count=n_segments,
                ),
                "no_speech_prob": np.fromiter(
                    (None if s.no_speech_prob is None else float(s.no_speech_prob) for s in segments),
                    dtype=object,
                    count=n_segments,
                ),
            }
        )
        segment_records = asr_df_raw.to_dict(orient="records")
    else:
        segment_records = []

    return {
        "backend": "faster-whisper",
        "backend_label": f"faster-whisper {asr_model} int8",
        "language": getattr(info, "language", None),
        "language_probability": getattr(info, "language_probability", None),
        "segment_records": segment_records,
        "model": f"faster-whisper-{asr_model}-int8",
    }


def transcribe_with_faster_whisper(audio_path, staging_root, video_id):
    result_path = Path(tempfile.mkdtemp(prefix=f"fw_asr_{video_id}_", dir=str(staging_root))) / "faster_whisper_asr.json"
    run(
        [
            sys.executable,
            str(Path(__file__).resolve()),
            "--internal-audio-path",
            str(audio_path),
            "--internal-asr-json-output",
            str(result_path),
            "--internal-asr-model",
            "small",
        ]
    )
    if not result_path.exists():
        raise FileNotFoundError(f"faster-whisper subprocess output not found: {result_path}")
    return json.loads(result_path.read_text())


def transcribe_with_whisper_cli(audio_path, whisper_model, staging_root, video_id):
    whisper_bin = shutil.which("whisper")
    if not whisper_bin:
        raise RuntimeError("whisper CLI not found in PATH")

    cli_output_dir = Path(tempfile.mkdtemp(prefix=f"whisper_cli_{video_id}_", dir=str(staging_root)))
    run(
        [
            whisper_bin,
            str(audio_path),
            "--model",
            whisper_model,
            "--device",
            "cpu",
            "--output_dir",
            str(cli_output_dir),
            "--output_format",
            "json",
            "--verbose",
            "False",
            "--fp16",
            "False",
        ]
    )

    json_path = cli_output_dir / f"{audio_path.stem}.json"
    if not json_path.exists():
        raise FileNotFoundError(f"whisper CLI output not found: {json_path}")

    whisper_obj = json.loads(json_path.read_text())
    segment_records = []
    for idx, segment in enumerate(whisper_obj.get("segments") or [], start=1):
        segment_records.append(
            {
                "id": segment.get("id", idx),
                "start": round(float(segment.get("start") or 0.0), 3),
                "end": round(float(segment.get("end") or 0.0), 3),
                "text": to_simplified_text((segment.get("text") or "").strip()),
                "avg_logprob": segment.get("avg_logprob"),
                "no_speech_prob": segment.get("no_speech_prob"),
            }
        )

    return {
        "backend": "whisper-cli",
        "backend_label": f"whisper CLI {whisper_model}",
        "language": whisper_obj.get("language"),
        "language_probability": None,
        "segment_records": segment_records,
        "model": f"openai-whisper-{whisper_model}-cli",
    }


def transcribe_audio(audio_path, asr_backend, whisper_cli_model, staging_root, video_id):
    backend_order = ["faster-whisper", "whisper-cli"] if asr_backend == "auto" else [asr_backend]
    attempt_errors = []

    for backend_name in backend_order:
        try:
            if backend_name == "faster-whisper":
                result = transcribe_with_faster_whisper(audio_path, staging_root, video_id)
            elif backend_name == "whisper-cli":
                result = transcribe_with_whisper_cli(audio_path, whisper_cli_model, staging_root, video_id)
            else:
                raise RuntimeError(f"Unsupported ASR backend: {backend_name}")
            result["attempt_errors"] = attempt_errors
            return result
        except Exception as exc:
            error_text = str(exc).strip() or repr(exc)
            attempt_errors.append({"backend": backend_name, "error": error_text})
            print(f"ASR backend failed ({backend_name}): {error_text.splitlines()[-1]}", flush=True)
            if asr_backend != "auto":
                raise

    raise RuntimeError(
        "All ASR backends failed: "
        + "; ".join(f"{item['backend']}: {item['error']}" for item in attempt_errors)
    )


def _json_unescape(s):
    if s is None:
        return None
    try:
        return json.loads(f"\"{s}\"")
    except Exception:
        return s


def _extract_json_string(blob, key):
    m = re.search(rf'"{re.escape(key)}"\s*:\s*"((?:\\.|[^"\\])*)"', blob)
    if not m:
        return None
    return _json_unescape(m.group(1))


def _extract_json_int(blob, key):
    m = re.search(rf'"{re.escape(key)}"\s*:\s*(\d+)', blob)
    if not m:
        return None
    try:
        return int(m.group(1))
    except Exception:
        return None


def _normalize_url_candidate(url):
    return (
        (url or "")
        .replace("\\u0026", "&")
        .replace("\\/", "/")
        .replace("\\u002F", "/")
        .strip()
    )


def _discover_douyin_raw_files(search_roots):
    files = []
    seen = set()
    patterns = ("**/douyin_*_raw.json", "**/*douyin*raw*.json")
    for root in search_roots:
        if not root:
            continue
        root = Path(root)
        if not root.exists():
            continue
        for pat in patterns:
            for p in root.glob(pat):
                if p.is_file() and p not in seen:
                    seen.add(p)
                    files.append(p)
    return files


def load_douyin_local_fallback(video_id, search_roots):
    for raw_path in _discover_douyin_raw_files(search_roots):
        try:
            text = raw_path.read_text(errors="ignore")
        except Exception:
            continue

        idx = text.find(f"\"aweme_id\": \"{video_id}\"")
        if idx < 0:
            continue

        window = text[max(0, idx - 2500): idx + 160000]

        play_urls = []
        for pat in (
            r'https://www\.douyin\.com/aweme/v1/play(?:/dash)?/\?[^"\\\s]+',
            r'https://[^"\\\s]+/mp4/main\.mp4[^"\\\s]*',
            r'https://[^"\\\s]*(?:douyinstatic\.com|zjcdn\.com)[^"\\\s]*',
        ):
            for m in re.finditer(pat, window):
                u = _normalize_url_candidate(m.group(0))
                if u and u not in play_urls:
                    play_urls.append(u)

        if not play_urls:
            continue

        desc = _extract_json_string(window, "desc")
        uploader = _extract_json_string(window, "nickname")
        uploader_id = _extract_json_string(window, "uid")
        channel_id = _extract_json_string(window, "sec_uid")
        if not channel_id:
            channel_id = uploader_id

        thumbnail = None
        tm = re.search(r'https://[^"\\\s]+\.(?:jpe?g|png)[^"\\\s]*', window)
        if tm:
            thumbnail = _normalize_url_candidate(tm.group(0))

        meta = {
            "id": video_id,
            "webpage_url": f"https://www.douyin.com/video/{video_id}",
            "description": desc,
            "timestamp": _extract_json_int(window, "create_time"),
            "uploader": uploader,
            "uploader_id": uploader_id,
            "channel": uploader,
            "channel_id": channel_id,
            "track": _extract_json_string(window, "title"),
            "view_count": _extract_json_int(window, "play_count"),
            "like_count": _extract_json_int(window, "digg_count"),
            "comment_count": _extract_json_int(window, "comment_count"),
            "repost_count": _extract_json_int(window, "share_count"),
            "save_count": _extract_json_int(window, "collect_count"),
            "thumbnail": thumbnail,
            "fallback_source": str(raw_path),
            "metadata_source": "local_douyin_raw",
        }
        return meta, play_urls, raw_path

    return None, [], None


def ffprobe_duration(path):
    proc = run([
        "ffprobe", "-v", "error", "-show_entries", "format=duration",
        "-of", "default=noprint_wrappers=1:nokey=1", str(path)
    ])
    return float((proc.stdout or "0").strip() or 0)


def has_video_stream(path):
    proc = subprocess.run(
        [
            "ffprobe", "-v", "error",
            "-select_streams", "v:0",
            "-show_entries", "stream=codec_type",
            "-of", "default=noprint_wrappers=1:nokey=1",
            str(path),
        ],
        text=True,
        stdout=subprocess.PIPE,
        stderr=subprocess.PIPE,
    )
    return proc.returncode == 0 and bool((proc.stdout or "").strip())


def format_ts(sec):
    sec = max(0.0, float(sec))
    m = int(sec // 60)
    s = sec - 60 * m
    return f"{m:02d}:{s:04.1f}"


def extract_frames(video_path, frames_dir, duration):
    if duration <= 0.5:
        times = [0.0]
    else:
        n = 12 if duration >= 24 else max(6, int(round(duration / 4)))
        n = max(2, n)
        times = [round(i * duration / (n - 1), 3) for i in range(n)]

    out_paths = []
    resolved_times = []
    max_seek = max(0.0, duration - 0.35)

    for idx, t in enumerate(times, start=1):
        t = min(float(t), max_seek)
        out = frames_dir / f"frame_{idx:03d}.jpg"
        attempt_times = []
        for candidate in (t, max(0.0, t - 0.8), max(0.0, t - 1.6)):
            candidate = round(candidate, 3)
            if candidate not in attempt_times:
                attempt_times.append(candidate)

        for attempt_t in attempt_times:
            if out.exists():
                out.unlink()
            run([
                "ffmpeg", "-y", "-hide_banner", "-loglevel", "error",
                "-ss", f"{attempt_t:.3f}", "-i", str(video_path), "-frames:v", "1", "-q:v", "2", str(out)
            ])
            if out.exists() and out.stat().st_size > 0:
                out_paths.append(out)
                resolved_times.append(attempt_t)
                break
        else:
            raise RuntimeError(
                f"Frame extraction failed for {out.name} near {t:.3f}s from {video_path}"
            )

    return resolved_times, out_paths


def detect_cuts(video_path, duration):
    scene_log = Path(f"/tmp/tiktok_scene_{video_path.stem}.log")
    if scene_log.exists():
        scene_log.unlink()

    run([
        "ffmpeg", "-y", "-hide_banner", "-loglevel", "error",
        "-i", str(video_path),
        "-filter_complex", f"select='gt(scene,0.10)',metadata=print:file={scene_log}",
        "-an", "-f", "null", "-"
    ])

    cut_times = []
    if scene_log.exists():
        text = scene_log.read_text(errors="ignore")
        for m in re.finditer(r"pts_time:([0-9.]+)", text):
            t = float(m.group(1))
            if 0.05 < t < max(0.05, duration - 0.05):
                cut_times.append(round(t, 6))

    cut_times = sorted(set(cut_times))

    dedup = []
    for t in cut_times:
        if not dedup or abs(t - dedup[-1]) >= 0.75:
            dedup.append(t)
    cut_times = dedup

    if not cut_times and duration > 4:
        anchors = 4 if duration >= 12 else 2
        cut_times = [round((i + 1) * duration / (anchors + 1), 6) for i in range(anchors)]

    return cut_times, scene_log


def _load_font(size):
    candidates = [
        "/Library/Fonts/Arial.ttf",
        "/System/Library/Fonts/Supplemental/Arial.ttf",
        "/System/Library/Fonts/Supplemental/Helvetica.ttc",
        "DejaVuSans.ttf",
    ]
    for fp in candidates:
        try:
            return ImageFont.truetype(fp, size)
        except Exception:
            continue
    return ImageFont.load_default()


def _save_storyboard(img, path, as_png=False):
    if as_png:
        img.save(path, format="PNG", optimize=True)
    else:
        img.save(path, format="JPEG", quality=98, subsampling=0, optimize=True)


def _render_storyboard(imgs, frame_times, out_plain, out_ts, tile_w, tile_h, cols, as_png=False):
    n = len(imgs)
    rows = math.ceil(n / cols)
    margin = max(10, int(min(tile_w, tile_h) * 0.03))
    top_pad = max(8, int(min(tile_w, tile_h) * 0.02))

    canvas_w = cols * tile_w + (cols + 1) * margin
    canvas_h = rows * tile_h + (rows + 1) * margin + top_pad

    plain = Image.new("RGB", (canvas_w, canvas_h), "white")
    ts = Image.new("RGB", (canvas_w, canvas_h), "white")
    draw = ImageDraw.Draw(ts, "RGBA")
    font = _load_font(max(18, int(min(tile_w, tile_h) * 0.07)))

    for idx, (img, t) in enumerate(zip(imgs, frame_times)):
        r = idx // cols
        c = idx % cols
        x = margin + c * (tile_w + margin)
        y = margin + r * (tile_h + margin) + top_pad
        fitted = ImageOps.fit(img, (tile_w, tile_h), method=Image.Resampling.LANCZOS)
        plain.paste(fitted, (x, y))
        ts.paste(fitted, (x, y))

        label = format_ts(t)
        bbox = draw.textbbox((0, 0), label, font=font)
        tw = bbox[2] - bbox[0]
        th = bbox[3] - bbox[1]
        lp = max(6, int(min(tile_w, tile_h) * 0.025))
        lx1 = x + lp
        ly1 = y + lp
        lx2 = lx1 + tw + 2 * lp
        ly2 = ly1 + th + 2 * lp
        draw.rounded_rectangle((lx1, ly1, lx2, ly2), radius=max(6, lp), fill=(0, 0, 0, 180))
        draw.text((lx1 + lp, ly1 + lp), label, fill="white", font=font)

    _save_storyboard(plain, out_plain, as_png=as_png)
    _save_storyboard(ts, out_ts, as_png=as_png)


def build_storyboards(frame_paths, frame_times, out_plain, out_ts, out_plain_hd, out_ts_hd):
    imgs = [Image.open(p).convert("RGB") for p in frame_paths]
    if not imgs:
        raise RuntimeError("No frames found for storyboard generation")

    w0, h0 = imgs[0].size
    is_vertical = h0 >= w0

    if is_vertical:
        std_tile_w, std_tile_h, std_cols = 270, 480, 4
        hd_tile_w, hd_tile_h, hd_cols = 360, 640, 3
    else:
        std_tile_w, std_tile_h, std_cols = 480, 270, 4
        hd_tile_w, hd_tile_h, hd_cols = 640, 360, 3

    _render_storyboard(
        imgs, frame_times, out_plain, out_ts,
        tile_w=std_tile_w, tile_h=std_tile_h, cols=std_cols, as_png=False
    )
    _render_storyboard(
        imgs, frame_times, out_plain_hd, out_ts_hd,
        tile_w=hd_tile_w, tile_h=hd_tile_h, cols=hd_cols, as_png=True
    )


# 1) metadata and folder setup
args = parse_args()
if args.internal_asr_json_output:
    asr_result = _run_internal_faster_whisper(
        Path(args.internal_audio_path),
        args.internal_asr_model,
        Path(args.internal_asr_json_output),
    )
    out_path = Path(args.internal_asr_json_output)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    out_path.write_text(json.dumps(asr_result, ensure_ascii=False, indent=2))
    raise SystemExit(0)

URL = args.url
TEMPLATE_PATH = Path(args.template_path)
OUTPUT_ROOT = Path(args.output_root)

video_id = parse_video_id(URL)
final_out_dir = OUTPUT_ROOT / output_folder_name(URL, video_id)
staging_root = Path(tempfile.mkdtemp(prefix=f"mcp_stage_{video_id}_", dir="/tmp"))
out_dir = staging_root / output_folder_name(URL, video_id)
frames_dir = out_dir / "frames"
out_dir.mkdir(parents=True, exist_ok=True)
run_succeeded = False


def _cleanup_staging_dir():
    if run_succeeded:
        return
    try:
        if staging_root.exists():
            shutil.rmtree(staging_root)
    except Exception:
        pass


atexit.register(_cleanup_staging_dir)
atexit.register(lambda: cleanup_tmp_media(video_id))

# Remove stale outputs from prior runs
if frames_dir.exists():
    shutil.rmtree(frames_dir)
frames_dir.mkdir(parents=True, exist_ok=True)
for stale in [
    "meta.json", "storyboard.jpg", "storyboard_ts.jpg", "storyboard_hd.png", "storyboard_ts_hd.png", "audio.wav", "asr.json",
    "cut_pts_tokens.txt", "cut_pts_tokens_low.txt", "cut_times.txt", "cut_times_raw.txt",
    "frame_times_raw.txt", "yt_scene.log", f"Climate_SVP_MCP_{video_id}_transcript_v3_filled.xlsx",
    "video_url.txt"
]:
    p = out_dir / stale
    if p.exists():
        p.unlink()

# Safety: never keep source/temporary video files in output folders.
remove_output_video_files(out_dir)

(out_dir / "video_url.txt").write_text(URL + "\n")

fallback_play_urls = []
fallback_raw_json_path = None
try:
    meta_proc = run(["yt-dlp", "-J", "--no-playlist", URL])
    meta = json.loads(meta_proc.stdout)
except Exception as e:
    if not is_douyin_video_url(URL):
        remove_output_video_files(out_dir)
        raise

    search_roots = [
        OUTPUT_ROOT,
        OUTPUT_ROOT.parent,
        OUTPUT_ROOT.parent.parent if OUTPUT_ROOT.parent else None,
        Path("/Users/XM/Desktop/Data/DY data"),
    ]
    meta, fallback_play_urls, fallback_raw_json_path = load_douyin_local_fallback(video_id, search_roots)
    if not meta:
        remove_output_video_files(out_dir)
        raise RuntimeError(
            f"yt-dlp metadata failed and no local Douyin fallback was found for {video_id}: {e}"
        )
    print(f"Using local Douyin fallback metadata for {video_id}", flush=True)

(out_dir / "meta.json").write_text(json.dumps(meta, ensure_ascii=False, indent=2))
publish_douyin_raw_json(fallback_raw_json_path, out_dir)

# 2) temp video download to /tmp only
cleanup_tmp_media(video_id)
manual_seed_path = Path(f"/tmp/dy_manual_{video_id}_source.mp4")
if manual_seed_path.exists():
    manual_dst = Path(f"/tmp/tiktok_{video_id}_manual.mp4")
    if not has_video_stream(manual_seed_path):
        manual_dst = Path(f"/tmp/tiktok_{video_id}_manual.m4a")
    shutil.copy2(manual_seed_path, manual_dst)

download_cmd = [
    "yt-dlp", "--no-playlist", "--force-overwrites",
    "-f", "bv*+ba/b", "--merge-output-format", "mp4",
    "-o", f"/tmp/tiktok_{video_id}.%(ext)s", URL
]
download_err = None
try:
    run(download_cmd)
except Exception as e:
    download_err = e

if download_err and fallback_play_urls:
    print(f"yt-dlp download failed, trying {len(fallback_play_urls)} local fallback play URLs", flush=True)
    for i, play_url in enumerate(fallback_play_urls, start=1):
        fb_tmp = Path(f"/tmp/tiktok_{video_id}_fb_{i:03d}.mp4")
        if fb_tmp.exists():
            try:
                fb_tmp.unlink()
            except Exception:
                pass
        run(["curl", "-L", "--fail", "-o", str(fb_tmp), play_url], check=False)
        if fb_tmp.exists() and fb_tmp.stat().st_size > 0 and has_video_stream(fb_tmp):
            print(f"Using fallback downloaded source: {fb_tmp}", flush=True)
            download_err = None
            break

if download_err and not fallback_play_urls:
    remove_output_video_files(out_dir)
    raise download_err

tmp_patterns = (f"tiktok_{video_id}.*", f"tiktok_{video_id}_*")
tmp_candidates = [
    p
    for pattern in tmp_patterns
    for p in Path("/tmp").glob(pattern)
    if p.suffix.lower() in {".mp4", ".mkv", ".webm", ".mov"} and p.is_file() and has_video_stream(p)
]
if not tmp_candidates:
    audio_candidates = [
        p
        for pattern in tmp_patterns
        for p in Path("/tmp").glob(pattern)
        if p.suffix.lower() in {".m4a", ".aac", ".mp3", ".wav", ".ogg", ".opus", ".weba", ".webm"} and p.is_file()
    ]
    if audio_candidates:
        audio_src = max(audio_candidates, key=lambda p: p.stat().st_size)
        thumb_url = meta.get("thumbnail")
        if not thumb_url:
            thumbs = meta.get("thumbnails") or []
            for t in thumbs:
                u = (t or {}).get("url")
                if u:
                    thumb_url = u
                    break

        thumb_path = Path(f"/tmp/tiktok_{video_id}_thumb.jpg")
        synth_video = Path(f"/tmp/tiktok_{video_id}_synth.mp4")
        for p in (thumb_path, synth_video):
            if p.exists():
                try:
                    p.unlink()
                except Exception:
                    pass

        have_thumb = False
        if thumb_url:
            run(["curl", "-L", "-o", str(thumb_path), thumb_url], check=False)
            have_thumb = thumb_path.exists() and thumb_path.stat().st_size > 0

        if have_thumb:
            run([
                "ffmpeg", "-y", "-hide_banner", "-loglevel", "error",
                "-loop", "1", "-i", str(thumb_path),
                "-i", str(audio_src),
                "-c:v", "libx264",
                "-vf", "scale=720:1280:force_original_aspect_ratio=decrease,pad=720:1280:(ow-iw)/2:(oh-ih)/2,format=yuv420p",
                "-pix_fmt", "yuv420p",
                "-c:a", "aac",
                "-shortest",
                str(synth_video),
            ])
        else:
            run([
                "ffmpeg", "-y", "-hide_banner", "-loglevel", "error",
                "-f", "lavfi", "-i", "color=c=black:s=720x1280:r=25",
                "-i", str(audio_src),
                "-c:v", "libx264",
                "-pix_fmt", "yuv420p",
                "-c:a", "aac",
                "-shortest",
                str(synth_video),
            ])

        if synth_video.exists() and synth_video.stat().st_size > 0 and has_video_stream(synth_video):
            tmp_candidates = [synth_video]

if not tmp_candidates:
    remove_output_video_files(out_dir)
    if download_err:
        raise download_err
    raise RuntimeError("No temporary video downloaded to /tmp (including fallback)")
video_path = max(tmp_candidates, key=lambda p: p.stat().st_size)
print(f"Using temporary source: {video_path}", flush=True)

# 3) frames + storyboards
duration = ffprobe_duration(video_path)
resolved_whisper_cli_model, whisper_cli_model_note = resolve_whisper_cli_model(
    requested_model=args.whisper_cli_model,
    asr_backend=args.asr_backend,
    duration_sec=duration,
)
if whisper_cli_model_note:
    print(whisper_cli_model_note, flush=True)
frame_times, frame_paths = extract_frames(video_path, frames_dir, duration)
(out_dir / "frame_times_raw.txt").write_text(
    ",".join(f"{t:.3f}" for t in frame_times)
)
build_storyboards(
    frame_paths, frame_times,
    out_plain=out_dir / "storyboard.jpg",
    out_ts=out_dir / "storyboard_ts.jpg",
    out_plain_hd=out_dir / "storyboard_hd.png",
    out_ts_hd=out_dir / "storyboard_ts_hd.png",
)
if not args.keep_jpg_storyboards:
    for legacy_storyboard in (out_dir / "storyboard.jpg", out_dir / "storyboard_ts.jpg"):
        if legacy_storyboard.exists():
            legacy_storyboard.unlink()

# 4) cuts
cut_times, scene_log = detect_cuts(video_path, duration)
if scene_log.exists():
    shutil.copy2(scene_log, out_dir / "yt_scene.log")

(out_dir / "cut_times_raw.txt").write_text(
    "\n".join(f"{t:.6f}" for t in cut_times) + ("\n" if cut_times else "")
)
(out_dir / "cut_times.txt").write_text(
    "\n".join(f"{t:.6f}" for t in cut_times) + ("\n" if cut_times else "")
)
(out_dir / "cut_pts_tokens.txt").write_text(
    "\n".join(f"pts_time={t:.6f}" for t in cut_times) + ("\n" if cut_times else "")
)
(out_dir / "cut_pts_tokens_low.txt").write_text(
    "\n".join(f"pts_time={t:.6f}" for t in cut_times[: max(1, min(8, len(cut_times)) )]) + ("\n" if cut_times else "")
)

# 5) audio + asr
audio_path = out_dir / "audio.wav"
audio_fallback_reason = None
try:
    run([
        "ffmpeg", "-y", "-hide_banner", "-loglevel", "error",
        "-i", str(video_path), "-vn", "-ac", "1", "-ar", "16000", str(audio_path)
    ])
except RuntimeError as e:
    err = str(e)
    if ("does not contain any stream" in err) or ("matches no streams" in err):
        audio_fallback_reason = "No audio stream in source; generated silent WAV fallback."
        silent_dur = max(1.0, float(duration))
        run([
            "ffmpeg", "-y", "-hide_banner", "-loglevel", "error",
            "-f", "lavfi", "-i", "anullsrc=channel_layout=mono:sample_rate=16000",
            "-t", f"{silent_dur:.3f}", "-acodec", "pcm_s16le", str(audio_path)
        ])
    else:
        raise

asr_result = transcribe_audio(
    audio_path,
    asr_backend=args.asr_backend,
    whisper_cli_model=resolved_whisper_cli_model,
    staging_root=staging_root,
    video_id=video_id,
)
asr_segments, transcript_full, duration_after_vad = summarize_asr_segments(
    asr_result.get("segment_records") or []
)

asr_obj = {
    "language": asr_result.get("language"),
    "language_probability": asr_result.get("language_probability"),
    "duration": round(duration, 3),
    "duration_after_vad": duration_after_vad,
    "segment_count": len(asr_segments),
    "segments": asr_segments,
    "transcript_full": transcript_full,
    "model": asr_result.get("model"),
    "asr_backend": asr_result.get("backend"),
}
if audio_fallback_reason:
    asr_obj["audio_note"] = audio_fallback_reason
if asr_result.get("attempt_errors"):
    asr_obj["backend_attempt_errors"] = asr_result.get("attempt_errors")
asr_obj = to_simplified_obj(asr_obj)
(out_dir / "asr.json").write_text(json.dumps(asr_obj, ensure_ascii=False, indent=2))

# 6) lightweight OCR from representative frames via tesseract CLI
ocr_events = []
seen_text = set()
for idx, (fp, t) in enumerate(zip(frame_paths, frame_times), start=1):
    proc = subprocess.run(
        ["tesseract", str(fp), "stdout", "--psm", "6", "-l", "eng"],
        stdout=subprocess.PIPE,
        stderr=subprocess.PIPE,
    )
    ocr_stdout = (proc.stdout or b"").decode("utf-8", errors="replace")
    txt = to_simplified_text(re.sub(r"\s+", " ", ocr_stdout).strip())
    if len(txt) < 5 or not re.search(r"[A-Za-z]", txt):
        continue
    key = txt.lower()
    if key in seen_text:
        continue
    seen_text.add(key)
    ocr_events.append({
        "ocr_id": f"OCR{idx:03d}",
        "start_sec": round(max(0.0, t - 1.5), 3),
        "end_sec": round(min(duration, t + 1.5), 3),
        "text_content": txt[:600],
        "position": "mixed",
        "type": "subtitle/overlay",
        "language": "EN",
        "confidence": 0.72,
        "notes": "OCR from representative sampled frames via tesseract.",
    })

if not ocr_events:
    ocr_events.append({
        "ocr_id": "OCR001",
        "start_sec": 0,
        "end_sec": round(min(duration, 3.0), 3),
        "text_content": "No reliable OCR text detected from representative frames.",
        "position": "NA",
        "type": "NA",
        "language": "EN",
        "confidence": 0.4,
        "notes": "Fallback record to keep OCR_Events populated.",
    })

# 7) populate workbook
filled_xlsx = out_dir / f"Climate_SVP_MCP_{video_id}_transcript_v3_filled.xlsx"
shutil.copy2(TEMPLATE_PATH, filled_xlsx)
wb = openpyxl.load_workbook(filled_xlsx)

vs = wb["Video_Summary"]
es = wb["Engagement_Snapshots"]
ts_ws = wb["Transcript_Segments"]
ocr_ws = wb["OCR_Events"]
scenes_ws = wb["Scenes"]
cuts_ws = wb["Cuts"]
summary_ws = wb["Summary_Card"]


def map_headers(ws):
    m = {}
    for c in range(1, ws.max_column + 1):
        h = ws.cell(1, c).value
        if h is not None:
            m[str(h)] = c
    return m


def clear_sheet_data_values(ws, columns):
    col_set = set(columns)
    if not col_set:
        return

    # Fast path: only touch populated cell objects.
    cell_store = getattr(ws, "_cells", None)
    if isinstance(cell_store, dict):
        for (row_idx, col_idx), cell in cell_store.items():
            if row_idx >= 2 and col_idx in col_set and cell.value is not None:
                cell.value = None
        return

    min_col = min(col_set)
    max_col = max(col_set)
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=min_col, max_col=max_col):
        for cell in row:
            if cell.column in col_set and cell.value is not None:
                cell.value = None


def write_dataframe_rows(ws, df, ws_columns):
    if df.empty:
        return
    values = df.where(pd.notna(df), None).to_numpy(dtype=object, copy=False)
    for row_idx, row_values in enumerate(values, start=2):
        for col_idx, value in zip(ws_columns, row_values):
            ws.cell(row_idx, int(col_idx), value=value)


vhm = map_headers(vs)


def set_vs(header, value):
    c = vhm.get(header)
    if c is not None:
        vs.cell(2, c, value=value)

collected_at = now_iso_utc()
upload_iso = iso_utc_from_ts(meta.get("timestamp"))
description = to_simplified_text((meta.get("description") or "").strip())
hashtags = parse_hashtags(description)
hashtag_join = ";".join(hashtags) if hashtags else None
asr_lang = (asr_obj.get("language") or "EN").upper()
asr_backend_used = asr_obj.get("asr_backend") or "unknown"
if asr_backend_used == "faster-whisper":
    transcript_source_text = "ASR (faster-whisper small int8, audio extracted from temporary /tmp cache)"
elif asr_backend_used == "whisper-cli":
    transcript_source_text = (
        f"ASR (whisper CLI {resolved_whisper_cli_model}, audio extracted from temporary /tmp cache)"
    )
else:
    transcript_source_text = "ASR (backend unknown, audio extracted from temporary /tmp cache)"
asr_backend_errors = asr_obj.get("backend_attempt_errors") or []

# Vectorized confidence proxy from avg_logprob.
asr_df = pd.DataFrame(asr_segments)
if asr_df.empty:
    asr_conf = np.array([], dtype=float)
else:
    if "avg_logprob" in asr_df.columns:
        lp_arr = pd.to_numeric(asr_df["avg_logprob"], errors="coerce").to_numpy(dtype=float, copy=False)
    else:
        lp_arr = np.full(len(asr_df), np.nan, dtype=float)
    asr_conf = np.clip(np.exp(lp_arr), 0.0, 1.0)
    asr_conf[~np.isfinite(asr_conf)] = np.nan

conf_mean = round(float(np.nanmean(asr_conf)), 3) if asr_conf.size and np.isfinite(asr_conf).any() else None

speech_seconds = asr_obj.get("duration_after_vad") or 0
transcript_available = 1 if transcript_full else 0
ocr_df = pd.DataFrame(ocr_events)
if ocr_df.empty:
    ocr_words_total = 0
else:
    if "text_content" in ocr_df.columns:
        ocr_text = ocr_df["text_content"].fillna("").astype(str)
    else:
        ocr_text = pd.Series([""] * len(ocr_df))
    ocr_words_total = int(
        ocr_text[~ocr_text.str.contains("No reliable OCR", regex=False)]
        .str.split()
        .str.len()
        .sum()
    )

is_science_explainer = bool(re.search(r"chaos|theory|curve|math|equation|diff", description.lower()))
has_climate = "climate" in description.lower() or "climate" in transcript_full.lower()

set_vs("platform", "TikTok")
set_vs("video_id", video_id)
set_vs("video_url", meta.get("webpage_url") or URL)
set_vs("creator_id", str(meta.get("uploader_id") or meta.get("channel_id") or ""))
set_vs("creator_username", meta.get("uploader") or meta.get("channel") or "")
set_vs("upload_datetime", upload_iso)
set_vs("collected_at", collected_at)
set_vs("duration_sec", round(duration, 3))
set_vs("language_primary", asr_lang)
set_vs("hashtags (semicolon)", hashtag_join)
set_vs("music_title", meta.get("track") or None)
set_vs("sound_type", "Speech + music/track" if transcript_available else "Music/ambient or low-speech")
set_vs("views", meta.get("view_count"))
set_vs("likes", meta.get("like_count"))
set_vs("comments", meta.get("comment_count"))
set_vs("shares", meta.get("repost_count"))
set_vs("saves", meta.get("save_count"))
set_vs("speech_seconds", round(float(speech_seconds), 3))
set_vs("transcript_available (0/1)", transcript_available)
set_vs("ocr_words_total", int(ocr_words_total))
set_vs("cut_count", len(cut_times))
set_vs("talking_head_ratio (0-1)", 0.35 if transcript_available else 0.1)
set_vs("data_visual_ratio (0-1)", 0.78 if is_science_explainer else 0.3)
set_vs("disaster_imagery_ratio (0-1)", 0.0)
set_vs("narrative_label", "Scientific explainer" if is_science_explainer else "Other/Unclear")
set_vs("entangled (0/1)", 1 if has_climate else 0)
set_vs("entanglement_score (0-1)", 0.68 if has_climate else 0.25)
set_vs(
    "entanglement_domains (semicolon)",
    "Climate;Mathematics;Complex systems" if has_climate else "Mathematics"
)
set_vs("problem_present (0/1)", 1 if has_climate else 0)
set_vs("cause_attribution", "Unspecified")
set_vs("solution_present (0/1)", 0)
set_vs("solution_type", "Unspecified")
set_vs("emotion_label", "Neutral")
desc_clean = re.sub(r"\s+", " ", description).strip()
desc_snippet = desc_clean[:220] if desc_clean else ""
transcript_snippet = transcript_full[:220] if transcript_full else ""

if desc_snippet:
    emotional_arc_text = f"Auto-derived from metadata/ASR: {desc_snippet}"
else:
    emotional_arc_text = "Auto-derived from metadata/ASR; manual review recommended."

if transcript_snippet:
    key_message_text = transcript_snippet
elif desc_snippet:
    key_message_text = desc_snippet
else:
    key_message_text = "No reliable speech/caption summary detected from this run."

set_vs(
    "emotional_arc (free text)",
    emotional_arc_text
)
set_vs(
    "key_message (free text)",
    key_message_text
)
set_vs("call_to_action_type", "Information/Explain")
set_vs("cta_text (free text)", "None")
set_vs("stance_score (-2..+2, optional)", 1 if has_climate else 0)
set_vs("confidence_overall (0-1)", 0.82 if transcript_available else 0.62)
set_vs("label_source", "Hybrid")
notes_text = (
    "Full rerun completed: metadata, frames, storyboard, ASR, OCR, scenes, and cuts. "
    "Temporary source video stored only in /tmp and removed after processing."
)
if audio_fallback_reason:
    notes_text += f" {audio_fallback_reason}"
if asr_backend_used == "whisper-cli":
    notes_text += f" ASR completed via whisper CLI fallback ({resolved_whisper_cli_model})."
    if whisper_cli_model_note:
        notes_text += f" {whisper_cli_model_note}"
if asr_backend_errors:
    notes_text += " ASR backend failure(s): " + "; ".join(
        f"{item.get('backend')}: {str(item.get('error') or '').splitlines()[-1][:180]}"
        for item in asr_backend_errors
    )
set_vs("notes", notes_text)
set_vs("transcript_status", "Available" if transcript_available else "Unavailable")
set_vs("transcript_source", transcript_source_text)
set_vs("transcript_language", asr_lang)
set_vs("transcript_text (full)", transcript_full if transcript_full else None)
set_vs("transcript_snippet (<=500 chars)", transcript_full[:500] if transcript_full else None)
set_vs("transcript_char_count", len(transcript_full) if transcript_full else 0)
set_vs("transcript_confidence_mean (0-1)", conf_mean)

# Engagement_Snapshots
es.cell(2, 1, value=collected_at)
es.cell(2, 2, value="TikTok")
es.cell(2, 3, value=video_id)
es.cell(2, 4, value=meta.get("view_count"))
es.cell(2, 5, value=meta.get("like_count"))
es.cell(2, 6, value=meta.get("comment_count"))
es.cell(2, 7, value=meta.get("repost_count"))
es.cell(2, 8, value=meta.get("save_count"))
es.cell(2, 9, value=f"Snapshot collected on {collected_at} during full MCP rerun.")

# Transcript_Segments
if asr_df.empty:
    ts_rows_df = pd.DataFrame(
        [[video_id, "T001", 0, 0, "NA", "No transcript available.", asr_lang, 0.0, "ASR unavailable."]],
        columns=[
            "video_id", "segment_id", "start_sec", "end_sec", "speaker",
            "text", "language", "confidence", "notes",
        ],
    )
else:
    n_asr = len(asr_df)
    ts_rows_df = pd.DataFrame({
        "video_id": np.repeat(video_id, n_asr),
        "segment_id": np.char.mod("T%03d", np.arange(1, n_asr + 1)),
        "start_sec": asr_df.get("start", pd.Series([None] * n_asr)),
        "end_sec": asr_df.get("end", pd.Series([None] * n_asr)),
        "speaker": np.repeat("Speaker_1", n_asr),
        "text": asr_df.get("text", pd.Series([""] * n_asr)),
        "language": np.repeat(asr_lang, n_asr),
        "confidence": np.round(asr_conf, 3),
        "notes": np.repeat("ASR transcript from extracted audio.", n_asr),
    })

clear_sheet_data_values(ts_ws, range(1, 10))
write_dataframe_rows(ts_ws, ts_rows_df, [1, 2, 3, 4, 5, 6, 7, 8, 9])

# OCR_Events
ocr_rows_df = ocr_df.reindex(
    columns=[
        "ocr_id", "start_sec", "end_sec", "text_content", "position",
        "type", "language", "confidence", "notes",
    ]
).copy()
ocr_rows_df.insert(0, "video_id", np.repeat(video_id, len(ocr_rows_df)))

clear_sheet_data_values(ocr_ws, range(1, 11))
write_dataframe_rows(ocr_ws, ocr_rows_df, [1, 2, 3, 4, 5, 6, 7, 8, 9, 10])

# Cuts
cut_arr = np.asarray(cut_times, dtype=float)
if cut_arr.size:
    n_cuts = cut_arr.size
    cuts_rows_df = pd.DataFrame({
        "video_id": np.repeat(video_id, n_cuts),
        "cut_id": np.char.mod("C%02d", np.arange(1, n_cuts + 1)),
        "cut_time": cut_arr,
        "cut_type": np.repeat("Hard cut / shot switch", n_cuts),
        "notes": np.repeat("Scene-change timestamp from ffmpeg detector on temporary cached source.", n_cuts),
    })
else:
    cuts_rows_df = pd.DataFrame(columns=["video_id", "cut_id", "cut_time", "cut_type", "notes"])

clear_sheet_data_values(cuts_ws, range(1, 6))
write_dataframe_rows(cuts_ws, cuts_rows_df, [1, 2, 3, 4, 5])

# Scenes
boundaries = [0.0] + [float(t) for t in cut_arr.tolist() if 0.0 < t < duration] + [duration]
b = []
for x in sorted(boundaries):
    if not b or (x - b[-1]) >= 1.0:
        b.append(round(x, 3))
if b[0] != 0.0:
    b.insert(0, 0.0)
if b[-1] < duration:
    b.append(round(duration, 3))

if len(b) - 1 > 8:
    idxs = sorted(set(int(round(i * (len(b) - 1) / 8)) for i in range(9)))
    b = [b[i] for i in idxs]
    b[0] = 0.0
    b[-1] = round(duration, 3)

if len(b) - 1 < 3 and duration > 9:
    b = [0.0, round(duration / 3, 3), round(2 * duration / 3, 3), round(duration, 3)]

b = np.asarray(b, dtype=float)

scene_start = b[:-1]
scene_end = b[1:]
valid_scene = scene_end > scene_start
scene_start = scene_start[valid_scene]
scene_end = scene_end[valid_scene]

default_key_visual = "Representative visual segment from extracted frame samples."
scene_key_visual = np.full(scene_start.shape[0], default_key_visual, dtype=object)
if scene_start.size and not ocr_df.empty:
    ocr_match_df = ocr_df.reindex(columns=["start_sec", "end_sec", "text_content"]).copy()
    o_start = pd.to_numeric(ocr_match_df["start_sec"], errors="coerce").to_numpy(dtype=float, copy=False)
    o_end = pd.to_numeric(ocr_match_df["end_sec"], errors="coerce").to_numpy(dtype=float, copy=False)
    o_text = ocr_match_df["text_content"].fillna("").astype(str).str.slice(0, 180).to_numpy(dtype=object, copy=False)
    valid_ocr = np.isfinite(o_start) & np.isfinite(o_end)
    o_start = o_start[valid_ocr]
    o_end = o_end[valid_ocr]
    o_text = o_text[valid_ocr]
    if o_start.size:
        overlap = (o_start[None, :] <= scene_end[:, None]) & (o_end[None, :] >= scene_start[:, None])
        has_overlap = overlap.any(axis=1)
        first_match_idx = overlap.argmax(axis=1)
        scene_key_visual[has_overlap] = o_text[first_match_idx[has_overlap]]

n_scenes = scene_start.size
if n_scenes:
    scenes_rows_df = pd.DataFrame({
        "video_id": np.repeat(video_id, n_scenes),
        "scene_id": np.char.mod("S%d", np.arange(1, n_scenes + 1)),
        "start_sec": np.round(scene_start, 3),
        "end_sec": np.round(scene_end, 3),
        "source": np.repeat("Auto-derived from representative frames", n_scenes),
        "key_visual": scene_key_visual,
        "inference": np.repeat("Auto-inferred; manual review recommended", n_scenes),
        "progression": np.repeat("Sequential progression through animation/explainer content", n_scenes),
        "emotion": np.repeat("Neutral", n_scenes),
        "notes": np.repeat("Scene boundaries auto-generated from ffmpeg cuts with fallback smoothing.", n_scenes),
    })
else:
    scenes_rows_df = pd.DataFrame(
        columns=[
            "video_id", "scene_id", "start_sec", "end_sec", "source",
            "key_visual", "inference", "progression", "emotion", "notes",
        ]
    )

clear_sheet_data_values(scenes_ws, [1, 2, 3, 4, 6, 7, 8, 9, 10, 11])
write_dataframe_rows(scenes_ws, scenes_rows_df, [1, 2, 3, 4, 6, 7, 8, 9, 10, 11])

# Summary card selector
summary_ws["B3"] = video_id

normalize_workbook_to_simplified(wb)
wb.save(filled_xlsx)

# For Douyin URLs, remove legacy duplicate tiktok_<id>_mcp output folders.
if is_douyin_video_url(URL):
    legacy_dir = OUTPUT_ROOT / f"tiktok_{video_id}_mcp"
    if legacy_dir != out_dir and legacy_dir.exists():
        shutil.rmtree(legacy_dir)

# 8) cleanup temporary downloaded media from /tmp
cleanup_tmp_media(video_id)

# Final guardrail: output folders should never contain a cached source video.
remove_output_video_files(out_dir)

# 9) publish staged output only after successful completion.
OUTPUT_ROOT.mkdir(parents=True, exist_ok=True)
if final_out_dir.exists():
    shutil.rmtree(final_out_dir)
shutil.move(str(out_dir), str(final_out_dir))
run_succeeded = True
if staging_root.exists():
    shutil.rmtree(staging_root, ignore_errors=True)
out_dir = final_out_dir

print("DONE")
print(str(out_dir / "meta.json"))
print(str(out_dir / "frames"))
if (out_dir / "storyboard.jpg").exists():
    print(str(out_dir / "storyboard.jpg"))
if (out_dir / "storyboard_ts.jpg").exists():
    print(str(out_dir / "storyboard_ts.jpg"))
print(str(out_dir / "storyboard_hd.png"))
print(str(out_dir / "storyboard_ts_hd.png"))
print(str(out_dir / "audio.wav"))
print(str(out_dir / "asr.json"))
print(str(out_dir / f"Climate_SVP_MCP_{video_id}_transcript_v3_filled.xlsx"))
