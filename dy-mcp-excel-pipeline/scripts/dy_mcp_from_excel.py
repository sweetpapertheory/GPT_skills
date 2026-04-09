#!/usr/bin/env python3
import argparse
import re
import shutil
import subprocess
import sys
from pathlib import Path

from openpyxl import load_workbook


DEFAULT_PIPELINE_SCRIPT = (
    Path(__file__).resolve().parents[2]
    / "tiktok-mcp-excel-pipeline"
    / "scripts"
    / "tiktok_mcp_pipeline.py"
)
DEFAULT_TEMPLATE_PATH = Path("/Users/XM/Desktop/Climate_SVP_MCP_Template_transcript_v3.xlsx")
DEFAULT_OUTPUT_ROOT = Path("/Users/XM/Desktop/Data/DY data/DY Data analyais")


def parse_args():
    parser = argparse.ArgumentParser(
        description="Run DY MCP pipeline using URLs stored in an Excel sheet."
    )
    parser.add_argument(
        "--excel-path",
        required=True,
        help="Workbook path that contains Douyin/TikTok URLs.",
    )
    parser.add_argument(
        "--rows",
        required=True,
        help="Row selector. Examples: 101 or 101,102,110-115",
    )
    parser.add_argument(
        "--url-col",
        default="D",
        help="Column letter that stores URLs (default: D).",
    )
    parser.add_argument(
        "--sheet-name",
        default=None,
        help="Worksheet name. Defaults to active sheet.",
    )
    parser.add_argument(
        "--template-path",
        default=str(DEFAULT_TEMPLATE_PATH),
        help="Climate SVP MCP template workbook.",
    )
    parser.add_argument(
        "--output-root",
        default=str(DEFAULT_OUTPUT_ROOT),
        help="Output root folder.",
    )
    parser.add_argument(
        "--pipeline-script",
        default=str(DEFAULT_PIPELINE_SCRIPT),
        help="Underlying MCP pipeline script path.",
    )
    parser.add_argument(
        "--continue-on-error",
        action="store_true",
        help="Continue remaining rows after one row fails.",
    )
    parser.add_argument(
        "--keep-jpg-storyboards",
        action="store_true",
        help="Pass through to pipeline (default keeps only HD PNG storyboards).",
    )
    parser.add_argument(
        "--asr-backend",
        choices=["auto", "faster-whisper", "whisper-cli"],
        default="auto",
        help="ASR backend passed to the underlying pipeline (default: auto fallback).",
    )
    parser.add_argument(
        "--whisper-cli-model",
        default="turbo",
        help="Whisper CLI model name when CLI fallback is used (default: turbo).",
    )
    return parser.parse_args()


def parse_rows(rows_spec):
    rows = []
    for chunk in rows_spec.split(","):
        piece = chunk.strip()
        if not piece:
            continue
        if "-" in piece:
            left, right = piece.split("-", 1)
            start = int(left)
            end = int(right)
            if end < start:
                raise ValueError(f"Invalid descending range: {piece}")
            rows.extend(range(start, end + 1))
        else:
            rows.append(int(piece))
    deduped = []
    seen = set()
    for row in rows:
        if row not in seen:
            seen.add(row)
            deduped.append(row)
    if not deduped:
        raise ValueError("No rows parsed from --rows")
    return deduped


def parse_video_id(url):
    m = re.search(r"/video/(\d+)", url or "")
    if not m:
        raise ValueError(f"Cannot parse video ID from URL: {url}")
    return m.group(1)


def is_douyin_video_url(url):
    return "douyin.com/video/" in (url or "")


def output_dir_for(url, video_id, output_root):
    prefix = "Douyin" if is_douyin_video_url(url) else "tiktok"
    return Path(output_root) / f"{prefix}_{video_id}_mcp"


def read_url_from_excel(excel_path, sheet_name, url_col, row):
    wb = load_workbook(excel_path, read_only=True, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb.active
    if ws is None:
        wb.close()
        raise ValueError("Workbook has no readable worksheet")
    value = ws[f"{url_col.upper()}{row}"].value
    wb.close()
    if value is None:
        raise ValueError(f"Cell {url_col.upper()}{row} is empty")
    url = str(value).strip()
    if not url:
        raise ValueError(f"Cell {url_col.upper()}{row} is blank")
    return url


def cleanup_partial_output(out_dir):
    if not out_dir.exists():
        return
    shutil.rmtree(out_dir)


def run_row(args, row):
    url = read_url_from_excel(args.excel_path, args.sheet_name, args.url_col, row)
    video_id = parse_video_id(url)
    out_dir = output_dir_for(url, video_id, args.output_root)
    print(f"[row {row}] URL: {url}", flush=True)
    print(f"[row {row}] target folder: {out_dir}", flush=True)

    cmd = [
        sys.executable,
        str(Path(args.pipeline_script)),
        "--url",
        url,
        "--template-path",
        str(Path(args.template_path)),
        "--output-root",
        str(Path(args.output_root)),
        "--asr-backend",
        args.asr_backend,
        "--whisper-cli-model",
        args.whisper_cli_model,
    ]
    if args.keep_jpg_storyboards:
        cmd.append("--keep-jpg-storyboards")

    print("+", " ".join(cmd), flush=True)
    out_dir_preexisting = out_dir.exists()

    try:
        subprocess.run(cmd, check=True)

        filled_xlsx = out_dir / f"Climate_SVP_MCP_{video_id}_transcript_v3_filled.xlsx"
        if not filled_xlsx.exists():
            raise RuntimeError(f"Filled workbook not found: {filled_xlsx}")
    except Exception:
        if (not out_dir_preexisting) and out_dir.exists():
            try:
                cleanup_partial_output(out_dir)
                print(f"[row {row}] removed partial output: {out_dir}", flush=True)
            except Exception as cleanup_err:
                print(
                    f"[row {row}] warning: failed to remove partial output {out_dir}: {cleanup_err}",
                    flush=True,
                )
        raise

    print(f"[row {row}] done: {filled_xlsx}", flush=True)
    return filled_xlsx


def main():
    args = parse_args()
    rows = parse_rows(args.rows)

    failures = []
    for row in rows:
        try:
            run_row(args, row)
        except Exception as e:
            failures.append((row, str(e)))
            print(f"[row {row}] failed: {e}", flush=True)
            if not args.continue_on_error:
                break

    if failures:
        print("\nFailures:", flush=True)
        for row, err in failures:
            print(f"- row {row}: {err}", flush=True)
        raise SystemExit(1)

    print(f"\nCompleted rows: {', '.join(str(r) for r in rows)}", flush=True)


if __name__ == "__main__":
    main()
